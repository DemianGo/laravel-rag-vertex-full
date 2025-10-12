#!/usr/bin/env python3
import sys
import os
import tempfile
import shutil
try:
    from openpyxl import load_workbook
    input_file = sys.argv[1]
    
    # Handle files without extension
    if not input_file.endswith(('.xlsx', '.xls')):
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        shutil.copy2(input_file, temp_file.name)
        input_file = temp_file.name
        delete_temp = True
    else:
        delete_temp = False
    
    wb = load_workbook(input_file, read_only=True, data_only=True)
    total_rows = sum(sheet.max_row for sheet in wb.worksheets if sheet.max_row)
    print(max(1, total_rows))
    
    if delete_temp and os.path.exists(input_file):
        os.unlink(input_file)
except Exception as e:
    sys.exit(1)