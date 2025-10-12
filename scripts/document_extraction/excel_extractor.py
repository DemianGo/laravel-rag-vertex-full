#!/usr/bin/env python3
"""
Excel Extractor Wrapper
Directly extracts Excel files, works with temp files without extensions.
"""

import sys
import os
import shutil

def extract_excel(file_path: str) -> str:
    """Extract text from Excel"""
    temp_link = None
    try:
        if not os.path.exists(file_path):
            return ''
        
        # openpyxl requires .xlsx extension, create temp link if needed
        if not file_path.endswith('.xlsx') and not file_path.endswith('.xls'):
            temp_link = file_path + '.xlsx'
            shutil.copy2(file_path, temp_link)
            file_to_read = temp_link
        else:
            file_to_read = file_path
        
        from openpyxl import load_workbook
        
        workbook = load_workbook(file_to_read, read_only=True, data_only=True)
        
        text_parts = []
        for sheet in workbook.worksheets:
            text_parts.append(f"=== Sheet: {sheet.title} ===")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = ' | '.join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text_parts.append(row_text)
        
        workbook.close()
        return '\n'.join(text_parts)
    except Exception as e:
        # Silent fail - return empty
        return ''
    finally:
        # Clean up temp file
        if temp_link and os.path.exists(temp_link):
            try:
                os.remove(temp_link)
            except:
                pass

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(1)
    
    file_path = sys.argv[1]
    text = extract_excel(file_path)
    print(text)

