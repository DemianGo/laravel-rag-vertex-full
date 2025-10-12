#!/usr/bin/env python3
"""
Excel Structured Extractor
Extracts both text AND structured JSON data from Excel files.
Maintains backward compatibility while adding structured data support.
"""

import sys
import os
import json
import shutil
from typing import Dict, List, Any, Optional
from datetime import datetime, date


def extract_excel_structured(file_path: str) -> Dict[str, Any]:
    """
    Extract both text and structured data from Excel.
    
    Returns:
        {
            "success": bool,
            "text": str,  # For RAG chunking
            "structured_data": {  # For precise queries
                "sheets": [
                    {
                        "name": str,
                        "headers": [str],
                        "rows": [
                            {column: value}
                        ],
                        "row_count": int,
                        "column_count": int
                    }
                ],
                "metadata": {
                    "total_sheets": int,
                    "total_rows": int,
                    "file_size": int
                }
            },
            "chunking_hints": {  # For intelligent chunking
                "chunk_by": "row",  # row, sheet, or auto
                "preserve_headers": True
            }
        }
    """
    temp_link = None
    try:
        if not os.path.exists(file_path):
            return {
                "success": False,
                "error": "File not found",
                "text": "",
                "structured_data": None
            }
        
        # openpyxl requires .xlsx extension
        if not file_path.endswith('.xlsx') and not file_path.endswith('.xls'):
            temp_link = file_path + '.xlsx'
            shutil.copy2(file_path, temp_link)
            file_to_read = temp_link
        else:
            file_to_read = file_path
        
        from openpyxl import load_workbook
        
        workbook = load_workbook(file_to_read, read_only=True, data_only=True)
        
        text_parts = []
        structured_sheets = []
        total_rows = 0
        
        for sheet in workbook.worksheets:
            sheet_name = sheet.title
            text_parts.append(f"=== Sheet: {sheet_name} ===")
            
            # Extract all rows
            all_rows = list(sheet.iter_rows(values_only=True))
            
            if not all_rows:
                continue
            
            # First row as headers
            headers = []
            for cell in all_rows[0]:
                if cell is not None:
                    headers.append(str(cell).strip())
                else:
                    headers.append(f"Column_{len(headers) + 1}")
            
            # Add headers to text
            text_parts.append(' | '.join(headers))
            
            # Process data rows
            structured_rows = []
            for row_idx, row in enumerate(all_rows[1:], start=2):
                # Text representation (for RAG)
                row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                if row_text.strip():
                    text_parts.append(row_text)
                
                # Structured representation (for queries)
                row_dict = {}
                has_data = False
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers):
                        header = headers[col_idx]
                        value = serialize_cell_value(cell)
                        row_dict[header] = value
                        if value is not None and value != '':
                            has_data = True
                
                if has_data:
                    structured_rows.append(row_dict)
                    total_rows += 1
            
            # Add sheet to structured data
            structured_sheets.append({
                "name": sheet_name,
                "headers": headers,
                "rows": structured_rows,
                "row_count": len(structured_rows),
                "column_count": len(headers)
            })
        
        workbook.close()
        
        # Build result
        result = {
            "success": True,
            "text": '\n'.join(text_parts),
            "structured_data": {
                "sheets": structured_sheets,
                "metadata": {
                    "total_sheets": len(structured_sheets),
                    "total_rows": total_rows,
                    "file_size": os.path.getsize(file_path)
                }
            },
            "chunking_hints": {
                "chunk_by": "row" if total_rows < 10000 else "auto",
                "preserve_headers": True
            }
        }
        
        return result
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "text": "",
            "structured_data": None
        }
    finally:
        # Clean up temp file
        if temp_link and os.path.exists(temp_link):
            try:
                os.remove(temp_link)
            except:
                pass


def serialize_cell_value(cell: Any) -> Any:
    """
    Convert cell value to JSON-serializable type.
    Preserves numeric types for calculations.
    """
    if cell is None:
        return None
    
    # Date/datetime
    if isinstance(cell, (datetime, date)):
        return cell.isoformat()
    
    # Numbers (keep as numbers for calculations)
    if isinstance(cell, (int, float)):
        return cell
    
    # Boolean
    if isinstance(cell, bool):
        return cell
    
    # Everything else as string
    return str(cell).strip()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "No file provided"}))
        sys.exit(1)
    
    file_path = sys.argv[1]
    result = extract_excel_structured(file_path)
    print(json.dumps(result, ensure_ascii=False, indent=2))

